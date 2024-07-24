import openpyxl

class LoadExcel:
    def __init__(self, filename):
        self.filename = filename

    def reader(self):
        try:
            wb_obj = openpyxl.load_workbook(self.filename)
            sheet_obj = wb_obj.active
            m_row = sheet_obj.max_row
            data = []

            for i in range(1, m_row+1):
                for j in range(1,7):
                    cell_obj = sheet_obj.cell(row=i, column=j)
                    print(cell_obj.value)
                    data.append(cell_obj.value)
            print(data)
        except Exception as e:
            print(f"Error reading the Excel file: {e}")

excel_reader = LoadExcel('user.xlsx')
excel_reader.reader()
print(excel_reader.reader())
